import openpyxl

wb = openpyxl.load_workbook('programs/assembly/tage_test_trace_gemini.xlsx')
sheet = wb.active

headers = [
    'Order', 'PC', 'Instruction', 'x1', 'x4', 'x5', 'x14', 'x15', 'x17',
    'FTB Entry 0 (Src-Tgt)', 'FTB Entry 1 (Src-Tgt)',
    'Pre-GHR', 'Post-GHR', 'TAGE Details',
    'Pre-PHR', 'Post-PHR', 'ITTAGE Details'
]

for c, h in enumerate(headers, 1):
    sheet.cell(1, c).value = h

ghr = 0
phr = 0

ftb0 = "EMPTY"
ftb1 = "EMPTY"

def parse_val(val):
    if val is None:
        return 0
    s = str(val).strip().lower()
    if '0x' in s:
        s = s.split('0x')[-1]
    elif 'x' in s:
        s = s.split('x')[-1]
    return int(s, 16) if s else 0

for r in range(2, sheet.max_row + 1):
    order = sheet.cell(r, 1).value
    pc = sheet.cell(r, 2).value
    inst = sheet.cell(r, 3).value
    
    if order is None or pc is None:
        continue
        
    pc_val = parse_val(pc)
    inst_str = str(inst).strip()
    
    pre_ghr_str = f"0b{ghr:b}"
    pre_phr_str = f"0b{phr:b}"
    
    post_ghr_str = pre_ghr_str
    post_phr_str = pre_phr_str
    
    tage_detail = "-"
    ittage_detail = "-"
    
    if inst_str.startswith('beq') or inst_str.startswith('bne'):
        x14_val = int(sheet.cell(r, 7).value) if sheet.cell(r, 7).value is not None else 0
        
        taken = False
        if inst_str.startswith('beq'):
            taken = (x14_val == 0)
        elif inst_str.startswith('bne'):
            taken = True
            
        taken_bit = 1 if taken else 0
        
        idx = (pc_val ^ ghr) & 0x3f
        tag = (pc_val ^ ghr) & 0x3f
        
        if taken:
            tage_detail = f"T0[{idx}], Tag=0x{tag:02X}, US=0, CTR=4"
        else:
            tage_detail = "No Update (Not Taken)" if inst_str.startswith('beq') else f"T0[{idx}], Tag=0x{tag:02X}, US=0, CTR=4"
            
        ghr = ((ghr << 1) | taken_bit) & 0xFFFFFFFFFFFFFFFF
        post_ghr_str = f"0b{ghr:b}"
        
        if pc_val == 0x10 and taken:
            ftb0 = "x10 - x18"
        elif pc_val == 0x3c and taken:
            ftb1 = "x3c - x08"
            
    elif inst_str.startswith('jalr'):
        x4_val = parse_val(sheet.cell(r, 5).value)
        target = x4_val
        idx = (pc_val ^ phr) & 0x3f
        tag = (pc_val ^ phr) & 0x3f
        ittage_detail = f"T0[{idx}], Tag=0x{tag:02X}, Target=0x{target:02X}, US=0"
        
        phr = ((phr << 4) | (target & 0xF)) & 0xFFFFFFFFFFFFFFFF
        post_phr_str = f"0b{phr:b}"
        ftb1 = f"x34 - x{target:02x}"
        
    elif inst_str.startswith('jal x4'):
        ftb0 = "x1c - x30"
    elif inst_str.startswith('jal x0, 16'):
        ftb1 = "x2c - x3c"
    elif inst_str.startswith('jal x0, 24'):
        ftb1 = "x24 - x3c"

    sheet.cell(r, 10).value = ftb0
    sheet.cell(r, 11).value = ftb1
    sheet.cell(r, 12).value = pre_ghr_str
    sheet.cell(r, 13).value = post_ghr_str
    sheet.cell(r, 14).value = tage_detail
    sheet.cell(r, 15).value = pre_phr_str
    sheet.cell(r, 16).value = post_phr_str
    sheet.cell(r, 17).value = ittage_detail

wb.save('programs/assembly/tage_test_trace_gemini.xlsx')
print("Successfully updated tage_test_trace_gemini.xlsx!")
