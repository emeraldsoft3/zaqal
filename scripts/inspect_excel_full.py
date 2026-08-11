import openpyxl
wb = openpyxl.load_workbook('programs/assembly/tage_test_trace_gemini.xlsx')
sheet = wb.active
headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
print("Headers:", headers)
for r in range(1, 40):
    row_vals = [sheet.cell(r, c).value for c in range(1, len(headers) + 1)]
    if any(row_vals):
        print(f"R{r:2d}: {row_vals}")
