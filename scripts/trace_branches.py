import openpyxl

wb = openpyxl.load_workbook('programs/assembly/tage_test_trace_gemini.xlsx')
sheet = wb.active

for r in range(2, sheet.max_row + 1):
    order = sheet.cell(r, 1).value
    pc = sheet.cell(r, 2).value
    inst = sheet.cell(r, 3).value
    x14 = sheet.cell(r, 7).value
    if order is not None:
        print(f"Row {r:2d} | Order {order:2d} | PC {pc:4s} | Inst: {inst:20s} | x14: {str(x14):5s}")
