import openpyxl

wb = openpyxl.load_workbook('tage_test_trace_gemini.xlsx')
sheet = wb.active

headers = [cell.value for cell in sheet[1]]
tage_col = headers.index('TAGE Details') if 'TAGE Details' in headers else -1
ittage_col = headers.index('ITTAGE Details') if 'ITTAGE Details' in headers else -1
id_col = headers.index('ID') if 'ID' in headers else -1
pc_col = headers.index('PC') if 'PC' in headers else -1
inst_col = headers.index('Inst') if 'Inst' in headers else -1
cycle_col = headers.index('Cycle') if 'Cycle' in headers else -1

print("--- TAGE PREDICTIONS ---")
tage_found = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    tage_val = row[tage_col]
    if tage_val and 'Tag=' in str(tage_val):
        print(f"ID: {row[id_col]}, PC: {row[pc_col]}, Inst: {row[inst_col]}, Details: {tage_val}, Cycle: {row[cycle_col] if cycle_col != -1 else 'N/A'}")
        tage_found += 1
        if tage_found >= 5: break

print("\n--- ITTAGE PREDICTIONS ---")
ittage_found = 0
for row in sheet.iter_rows(min_row=2, values_only=True):
    ittage_val = row[ittage_col]
    if ittage_val and 'Tag=' in str(ittage_val):
        print(f"ID: {row[id_col]}, PC: {row[pc_col]}, Inst: {row[inst_col]}, Details: {ittage_val}, Cycle: {row[cycle_col] if cycle_col != -1 else 'N/A'}")
        ittage_found += 1
        if ittage_found >= 5: break
