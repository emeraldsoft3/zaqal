import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Instruction definitions
# Program starts at PC = 0x80000000 (offset 0)
code = [
    ("addi x1, x0, 10", "00a00093"),     # 0x00
    ("addi x5, x0, 0", "00000293"),      # 0x04
    ("addi x5, x5, 6", "00628293"),      # 0x08
    ("andi x14, x5, 3", "0032f713"),     # 0x0c
    ("beq x14, x0, 8", "00070463"),      # 0x10
    ("addi x15, x0, 1", "00100793"),     # 0x14
    ("slli x17, x14, 2", "00271893"),    # 0x18
    ("jal x4, 20", "0140026f"),          # 0x1c
    ("addi x15, x0, 10", "00a00793"),    # 0x20
    ("jal x0, 24", "0180006f"),          # 0x24
    ("addi x15, x0, 20", "01400793"),    # 0x28
    ("jal x0, 16", "0100006f"),          # 0x2c
    ("add x4, x4, x17", "01120233"),     # 0x30
    ("jalr x1, x4, 0", "000200e7"),      # 0x34
    ("addi x1, x1, -1", "fff08093"),     # 0x38
    ("bne x1, x0, -52", "fc0096e3"),     # 0x3c
    ("addi x12, x0, 99", "06300613")     # 0x40
]

def fold_history(history_bits, length, fold_width):
    safe_len = min(length, len(history_bits))
    chunks = (safe_len + fold_width - 1) // fold_width
    result = 0
    for i in range(chunks):
        start = i * fold_width
        end = min((i + 1) * fold_width, safe_len)
        chunk_val = 0
        for j in range(start, end):
            if history_bits[j] == 1:
                chunk_val |= (1 << (j - start))
        result ^= chunk_val
    return result

def get_phr_bits(phr_val):
    bits = []
    for i in range(32):
        bits.append((phr_val >> i) & 1)
    return bits

def generate_golden_trace(limit=200):
    pc = 0x80000000
    regs = {i: 0 for i in range(32)}
    
    # Predictor configuration matching Zaqal frontend / trace_generator.js
    TAGE_TABLES = 4
    TAGE_HIST_LENS = [4, 12, 36, 108]
    TAGE_INDEX_WIDTH = 7
    TAGE_TAG_WIDTH = 8
    
    ITTAGE_TABLES = 4
    ITTAGE_HIST_LENS = [4, 12, 36, 108]
    ITTAGE_INDEX_WIDTH = 6
    ITTAGE_TAG_WIDTH = 8
    
    ghr = [0] * 128
    phr = 0
    
    # Initialize Predictor Tables
    tage_tables = []
    for t in range(TAGE_TABLES):
        table = []
        for _ in range(1 << TAGE_INDEX_WIDTH):
            table.append({'valid': False, 'tag': 0, 'ctr': 3, 'u': 0})
        tage_tables.append(table)
        
    ittage_tables = []
    for t in range(ITTAGE_TABLES):
        table = []
        for _ in range(1 << ITTAGE_INDEX_WIDTH):
            table.append({'valid': False, 'tag': 0, 'target': 0, 'u': 0})
        ittage_tables.append(table)
        
    ftb = {}
    
    trace_data = []
    step_count = 0
    
    while step_count < limit:
        offset = pc - 0x80000000
        if offset < 0 or (offset // 4) >= len(code):
            break
            
        branch_pc = pc
        insn_str, hex_val = code[offset // 4]
        op = insn_str.split(" ")[0].strip().lower()
        
        is_cond = op in ["beq", "bne", "blt", "bge", "bltu", "bgeu"]
        is_jalr = op == "jalr"
        
        # Snapshots before execution
        row_ghr = list(ghr)
        row_phr = phr
        
        # 1. Predictor Lookup
        provider_tage = -1
        tage_indices = []
        tage_tags = []
        if is_cond:
            aligned_pc = (branch_pc & ~31) & 0xffffffff
            for t in range(TAGE_TABLES):
                hist_len = TAGE_HIST_LENS[t]
                idx_fh = fold_history(ghr, hist_len, TAGE_INDEX_WIDTH)
                tag_fh = fold_history(ghr, hist_len, TAGE_TAG_WIDTH)
                idx = (aligned_pc ^ idx_fh) & 127
                tag = (aligned_pc ^ tag_fh) & 255
                tage_indices.append(idx)
                tage_tags.append(tag)
                
                entry = tage_tables[t][idx]
                if entry['valid'] and entry['tag'] == tag:
                    provider_tage = t
                    
        provider_ittage = -1
        ittage_indices = []
        ittage_tags = []
        if is_jalr:
            aligned_pc = (branch_pc & ~31) & 0xffffffff
            phr_bits = get_phr_bits(phr)
            for t in range(ITTAGE_TABLES):
                hist_len = ITTAGE_HIST_LENS[t]
                idx_fh = fold_history(phr_bits, hist_len, ITTAGE_INDEX_WIDTH)
                tag_fh = fold_history(phr_bits, hist_len, ITTAGE_TAG_WIDTH)
                idx = (aligned_pc ^ idx_fh) & 63
                tag = (aligned_pc ^ tag_fh) & 255
                ittage_indices.append(idx)
                ittage_tags.append(tag)
                
                entry = ittable_entry = ittage_tables[t][idx]
                if entry['valid'] and entry['tag'] == tag:
                    provider_ittage = t

        # 2. Execute Instruction
        next_pc = pc + 4
        if op == "addi":
            # parse: addi rd, rs1, imm
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            rs1 = int(tokens[1][1:])
            imm = int(tokens[2])
            if rd != 0:
                regs[rd] = (regs[rs1] + imm) & 0xffffffff
                if regs[rd] & 0x80000000:
                    regs[rd] -= 0x100000000
        elif op == "andi":
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            rs1 = int(tokens[1][1:])
            imm = int(tokens[2])
            if rd != 0:
                regs[rd] = regs[rs1] & imm
        elif op == "slli":
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            rs1 = int(tokens[1][1:])
            imm = int(tokens[2])
            if rd != 0:
                regs[rd] = (regs[rs1] << imm) & 0xffffffff
                if regs[rd] & 0x80000000:
                    regs[rd] -= 0x100000000
        elif op == "add":
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            rs1 = int(tokens[1][1:])
            rs2 = int(tokens[2][1:])
            if rd != 0:
                regs[rd] = (regs[rs1] + regs[rs2]) & 0xffffffff
                if regs[rd] & 0x80000000:
                    regs[rd] -= 0x100000000
        elif op == "jal":
            # jal rd, imm
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            imm = int(tokens[1])
            if rd != 0:
                regs[rd] = pc + 4
            next_pc = pc + imm
        elif op == "jalr":
            # jalr rd, rs1, imm
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rd = int(tokens[0][1:])
            rs1 = int(tokens[1][1:])
            imm = int(tokens[2])
            temp_pc = pc + 4
            next_pc = (regs[rs1] + imm) & 0xfffffffe
            if rd != 0:
                regs[rd] = temp_pc
        elif op == "beq":
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rs1 = int(tokens[0][1:])
            rs2 = int(tokens[1][1:])
            imm = int(tokens[2])
            if regs[rs1] == regs[rs2]:
                next_pc = pc + imm
        elif op == "bne":
            tokens = [t.strip() for t in insn_str.split(" ", 1)[1].split(",")]
            rs1 = int(tokens[0][1:])
            rs2 = int(tokens[1][1:])
            imm = int(tokens[2])
            if regs[rs1] != regs[rs2]:
                next_pc = pc + imm
                
        step_count += 1
        after_pc = next_pc
        pc = next_pc
        
        actual_taken = False
        if is_cond:
            actual_taken = after_pc != (branch_pc + 4)
            
        # FTB Update
        is_cfi = is_jalr or (op == "jal") or (is_cond and actual_taken)
        if is_cfi:
            ftb_idx = (branch_pc >> 5) & 0x3f
            ftb[ftb_idx] = f"x{branch_pc & 0xff:02x} - x{after_pc & 0xff:02x}"
            
        # TAGE Update
        tage_details = "-"
        if is_cond:
            pT = provider_tage
            pIdx = tage_indices[pT] if pT != -1 else -1
            pTag = tage_tags[pT] if pT != -1 else -1
            
            provider_pred = True # Initial base table state is 2 (Taken)
            alt_pred = False
            
            if pT != -1:
                provider_pred = tage_tables[pT][pIdx]['ctr'] >= 4
                
                altT = -1
                altIdx = -1
                for t in range(pT - 1, -1, -1):
                    idx = tage_indices[t]
                    tag = tage_tags[t]
                    if tage_tables[t][idx]['valid'] and tage_tables[t][idx]['tag'] == tag:
                        altT = t
                        altIdx = idx
                        break
                if altT != -1:
                    alt_pred = tage_tables[altT][altIdx]['ctr'] >= 4
                else:
                    alt_pred = True
                    
            mispredict = provider_pred != actual_taken
            update_table = -1
            update_idx = -1
            update_tag = -1
            
            # Allocation
            if mispredict:
                for t in range(pT + 1, TAGE_TABLES):
                    idx = tage_indices[t]
                    tag = tage_tags[t]
                    if tage_tables[t][idx]['u'] == 0:
                        tage_tables[t][idx]['valid'] = True
                        tage_tables[t][idx]['tag'] = tag;
                        tage_tables[t][idx]['ctr'] = 4 if actual_taken else 3
                        tage_tables[t][idx]['u'] = 0
                        update_table = t
                        update_idx = idx
                        update_tag = tag
                        break
                if update_table == -1:
                    for t in range(pT + 1, TAGE_TABLES):
                        idx = tage_indices[t]
                        if tage_tables[t][idx]['u'] > 0:
                            tage_tables[t][idx]['u'] -= 1
            # Update Provider
            if pT != -1:
                entry = tage_tables[pT][pIdx]
                if actual_taken:
                    entry['ctr'] = min(entry['ctr'] + 1, 7)
                else:
                    entry['ctr'] = max(entry['ctr'] - 1, 0)
                if provider_pred != alt_pred:
                    if provider_pred == actual_taken:
                        entry['u'] = min(entry['u'] + 1, 3)
                    else:
                        entry['u'] = max(entry['u'] - 1, 0)
                if not mispredict:
                    update_table = pT
                    update_idx = pIdx
                    update_tag = pTag
                    
            if update_table != -1:
                entry = tage_tables[update_table][update_idx]
                tage_details = f"T{update_table}[{update_idx}], Tag=0x{update_tag:02X}, US={entry['u']}, CTR={entry['ctr']}"
            elif mispredict:
                tage_details = "Decay (No Alloc)"
            else:
                tage_details = "No Update"
                
            # Shift into GHR
            ghr.insert(0, 1 if actual_taken else 0)
            ghr.pop()
            
        # ITTAGE Update
        ittage_details = "-"
        if is_jalr:
            pT = provider_ittage
            pIdx = ittage_indices[pT] if pT != -1 else -1
            pTag = ittage_tags[pT] if pT != -1 else -1
            
            provider_pred_target = ittage_tables[pT][pIdx]['target'] if pT != -1 else 0
            mispredict = (pT == -1) or (provider_pred_target != after_pc)
            
            update_table = -1
            update_idx = -1
            update_tag = -1
            
            if pT != -1:
                entry = ittage_tables[pT][pIdx]
                if provider_pred_target == after_pc:
                    entry['u'] = min(entry['u'] + 1, 3)
                else:
                    entry['u'] = max(entry['u'] - 1, 0)
                    
            # Allocation
            if mispredict:
                for t in range(pT + 1, ITTAGE_TABLES):
                    idx = ittage_indices[t]
                    tag = ittage_tags[t]
                    if ittage_tables[t][idx]['u'] == 0:
                        ittage_tables[t][idx]['valid'] = True
                        ittage_tables[t][idx]['tag'] = tag
                        ittage_tables[t][idx]['target'] = after_pc
                        ittage_tables[t][idx]['u'] = 0
                        update_table = t
                        update_idx = idx
                        update_tag = tag
                        break
                if update_table == -1:
                    for t in range(pT + 1, ITTAGE_TABLES):
                        idx = ittage_indices[t]
                        if ittage_tables[t][idx]['u'] > 0:
                            ittage_tables[t][idx]['u'] -= 1
            elif pT != -1:
                update_table = pT
                update_idx = pIdx
                update_tag = pTag
                
            if update_table != -1:
                entry = ittage_tables[update_table][update_idx]
                ittage_details = f"T{update_table}[{update_idx}], Tag=0x{update_tag:02X}, Target=0x{entry['target'] & 0xff:02X}, US={entry['u']}"
            elif mispredict:
                ittage_details = "Decay (No Alloc)"
            else:
                ittage_details = "No Update"
                
            # Shift into PHR
            phr = ((phr << 6) | ((after_pc >> 2) & 0x3f)) & 0xffffffff

        # Format registers
        def fmt_reg(val, hex_fmt):
            if hex_fmt:
                return "0" if val == 0 else f"0x{val & 0xffffffff:X}"
            return str(val)
            
        def fmt_ghr(ghr_arr):
            val = 0
            for i in range(32):
                if ghr_arr[i] == 1:
                    val |= (1 << i)
            shifted = (val << 1) & 0xffffffff
            return "0" if shifted == 0 else f"0b{shifted:b}"
            
        def fmt_phr(p_val):
            return "0" if p_val == 0 else f"0b{p_val & 0xffffffff:b}"

        trace_data.append({
            'order': step_count,
            'pc': f"x{branch_pc & 0xff:02x}",
            'instruction': insn_str,
            'hex': hex_val,
            'x1': fmt_reg(regs[1], True),
            'x4': fmt_reg(regs[4], True),
            'x5': fmt_reg(regs[5], False),
            'x14': fmt_reg(regs[14], False),
            'x15': fmt_reg(regs[15], False),
            'x17': fmt_reg(regs[17], False),
            'ftb0': ftb.get(0, "EMPTY"),
            'ftb1': ftb.get(1, "EMPTY"),
            'ghr': fmt_ghr(row_ghr),
            'tage': tage_details,
            'pre_phr': fmt_phr(row_phr),
            'post_phr': fmt_phr(phr),
            'ittage': ittage_details
        })
        
    return trace_data

def write_excel(trace):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zaqal BPU Golden Trace"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    headers = [
        "Order", "PC", "Instruction", "Hex",
        "x1", "x4", "x5", "x14", "x15", "x17",
        "FTB Entry 0 (Src-Tgt)", "FTB Entry 1 (Src-Tgt)",
        "GHR (TAGE index)", "TAGE Details",
        "Pre PHR", "Post PHR", "ITTAGE Details"
    ]
    
    # Theme palette: Sleek Dark / Platinum
    header_fill = PatternFill(start_color="18181C", end_color="18181C", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color='E5E5E5'),
        right=Side(style='thin', color='E5E5E5'),
        top=Side(style='thin', color='E5E5E5'),
        bottom=Side(style='thin', color='E5E5E5')
    )
    
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Add data
    for idx, row in enumerate(trace):
        vals = [
            row['order'], row['pc'], row['instruction'], row['hex'],
            row['x1'], row['x4'], row['x5'], row['x14'], row['x15'], row['x17'],
            row['ftb0'], row['ftb1'],
            row['ghr'], row['tage'],
            row['pre_phr'], row['post_phr'], row['ittage']
        ]
        ws.append(vals)
        r_idx = idx + 2
        
        # Apply zebra striping and styling
        row_fill = PatternFill(start_color="FFFFFF" if idx % 2 == 0 else "F9F9FB", end_color="FFFFFF" if idx % 2 == 0 else "F9F9FB", fill_type="solid")
        row_font = Font(name="Segoe UI", size=10, color="1E1E24")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = row_font
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 2, 4]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in range(5, 11):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    out_path = 'programs/assembly/tage_test_trace_gemini.xlsx'
    wb.save(out_path)
    print(f"Successfully generated independent Golden Trace XLSX: {out_path}")

def write_html(trace):
    rows_html = ""
    for idx, row in enumerate(trace):
        stripe = "stripe-even" if idx % 2 == 0 else "stripe-odd"
        
        # Highlight predictions
        tage_cls = "text-muted" if row['tage'] == "-" else ("text-success" if "T0" in row['tage'] else "text-warning")
        ittage_cls = "text-muted" if row['ittage'] == "-" else "text-info"
        
        rows_html += f"""
        <tr class="{stripe}">
            <td style="text-align: center; font-weight: bold; color: #a9a9b3;">{row['order']}</td>
            <td style="text-align: center; color: #52c41a; font-family: monospace;">{row['pc']}</td>
            <td style="text-align: left; font-weight: 500;">{row['instruction']}</td>
            <td style="text-align: center; font-family: monospace; color: #8c8c8c;">{row['hex']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x1']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x4']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x5']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x14']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x15']}</td>
            <td style="text-align: right; font-family: monospace;">{row['x17']}</td>
            <td style="text-align: left; font-family: monospace; font-size: 0.9em; color: #a9a9b3;">{row['ftb0']}</td>
            <td style="text-align: left; font-family: monospace; font-size: 0.9em; color: #a9a9b3;">{row['ftb1']}</td>
            <td style="text-align: left; font-family: monospace; font-size: 0.85em; max-width: 150px; overflow-wrap: break-word;">{row['ghr']}</td>
            <td style="text-align: left; font-weight: bold;" class="{tage_cls}">{row['tage']}</td>
            <td style="text-align: left; font-family: monospace; font-size: 0.85em; max-width: 150px; overflow-wrap: break-word;">{row['pre_phr']}</td>
            <td style="text-align: left; font-family: monospace; font-size: 0.85em; max-width: 150px; overflow-wrap: break-word;">{row['post_phr']}</td>
            <td style="text-align: left; font-weight: bold;" class="{ittage_cls}">{row['ittage']}</td>
        </tr>
        """
        
    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Zaqal Branch Predictor Golden Trace Dashboard</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-color: #0b0b0d;
            --card-bg: #141416;
            --border-color: #1f1f23;
            --text-main: #e4e4e7;
            --text-muted: #a1a1aa;
            --primary: #6366f1;
            --success: #10b981;
            --info: #06b6d4;
            --warning: #f59e0b;
        }}
        
        body {{
            background-color: var(--bg-color);
            color: var(--text-main);
            font-family: 'Inter', sans-serif;
            margin: 0;
            padding: 24px;
        }}
        
        .header {{
            margin-bottom: 24px;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 16px;
        }}
        
        .header h1 {{
            margin: 0;
            font-size: 1.8rem;
            font-weight: 700;
            letter-spacing: -0.025em;
        }}
        
        .header p {{
            margin: 8px 0 0 0;
            color: var(--text-muted);
            font-size: 0.95rem;
        }}
        
        .card {{
            background-color: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.9rem;
        }}
        
        th {{
            background-color: #18181c;
            color: #ffffff;
            font-weight: 600;
            padding: 12px 8px;
            border-bottom: 1px solid var(--border-color);
            font-size: 0.85rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}
        
        td {{
            padding: 10px 8px;
            border-bottom: 1px solid var(--border-color);
            color: #d1d1d6;
        }}
        
        .stripe-even {{
            background-color: #141416;
        }}
        
        .stripe-odd {{
            background-color: #1a1a1e;
        }}
        
        .text-muted {{ color: var(--text-muted); }}
        .text-success {{ color: var(--success); }}
        .text-info {{ color: var(--info); }}
        .text-warning {{ color: var(--warning); }}
        
        tr:hover {{
            background-color: #27272a !important;
        }}
        
        td {{
            white-space: nowrap;
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Zaqal BPU Golden Trace Dashboard</h1>
        <p>Pure architectural simulation of conditional branches and indirect jumps (GHR, PHR, TAGE, ITTAGE, FTB).</p>
    </div>
    
    <div class="card" style="overflow-x: auto;">
        <table>
            <thead>
                <tr>
                    <th>Order</th>
                    <th>PC</th>
                    <th>Instruction</th>
                    <th>Hex</th>
                    <th>x1</th>
                    <th>x4</th>
                    <th>x5</th>
                    <th>x14</th>
                    <th>x15</th>
                    <th>x17</th>
                    <th>FTB Entry 0</th>
                    <th>FTB Entry 1</th>
                    <th>GHR (TAGE idx)</th>
                    <th>TAGE Details</th>
                    <th>Pre PHR</th>
                    <th>Post PHR</th>
                    <th>ITTAGE Details</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
    </div>
</body>
</html>
"""
    out_path = 'programs/assembly/tage_test_trace_gemini.html'
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"Successfully generated independent Golden Trace HTML: {out_path}")

if __name__ == '__main__':
    trace = generate_golden_trace(limit=200)
    write_excel(trace)
    write_html(trace)
