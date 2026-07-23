import os
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

disasm = {
    0x00: "addi x1, x0, 10",
    0x04: "addi x5, x0, 0",
    0x08: "addi x5, x5, 6",
    0x0c: "andi x14, x5, 3",
    0x10: "beq x14, x0, 8",
    0x14: "addi x15, x0, 1",
    0x18: "slli x17, x14, 2",
    0x1c: "jal x4, 20",
    0x20: "addi x15, x0, 10",
    0x24: "jal x0, 24",
    0x28: "addi x15, x0, 20",
    0x2c: "jal x0, 16",
    0x30: "add x4, x4, x17",
    0x34: "jalr x1, x4, 0",
    0x38: "addi x1, x1, -1",
    0x3c: "bne x1, x0, -52",
    0x40: "addi x12, x0, 99"
}

def parse_vcd_and_log():
    vcd_path = 'programs/vcd/Lithium.vcd'
    log_path = 'sim_output2.txt'

    if not os.path.exists(vcd_path) or not os.path.exists(log_path):
        print("Error: Lithium.vcd or sim_output2.txt not found.")
        return None, None

    print("Step 1: Parsing VCD Header & Building Signal Index...")
    scope_stack = []
    symbol_to_signals = {}
    
    # Track paths we need
    ghr_symbol = None
    phr_symbol = None
    clock_symbol = None

    # TAGE symbols
    tage_symbols = []
    for t in range(4):
        tage_symbols.append({
            'tag_en': [], 'tag_addr': [], 'tag_data': [],
            'ctr_en': [], 'ctr_addr': [], 'ctr_data': [],
            'us': {},
            'update_valid': [], 'allocate': [],
            'hit': [], 'read_tag_addr': [], 'req_tag': []
        })

    # ITTAGE symbols
    ittage_symbols = []
    for t in range(4):
        ittage_symbols.append({
            'tag_en': [], 'tag_addr': [], 'tag_data': [],
            'tgt_en': [], 'tgt_addr': [], 'tgt_data': [],
            'us': {},
            'update_valid': [], 'allocate': [],
            'hit': [], 'read_tag_addr': [], 'req_tag': []
        })

    # Regfile physical reg symbols
    preg_symbols = {} # index -> symbol

    with open(vcd_path, 'r') as f:
        for line in f:
            stripped = line.strip()
            if not stripped:
                continue
            if stripped.startswith('$scope'):
                parts = stripped.split()
                scope_stack.append(parts[2])
            elif stripped.startswith('$upscope'):
                if scope_stack:
                    scope_stack.pop()
            elif stripped.startswith('$var'):
                parts = stripped.split()
                symbol = parts[3]
                name_idx = 4
                while name_idx < len(parts) and parts[name_idx] != '$end':
                    name_idx += 1
                name = parts[4]
                full_path = ".".join(scope_stack) + "." + name
                
                # Check TOP.clock
                if full_path.endswith(".clock") or full_path.endswith("_clock") or full_path == "TOP.clock":
                    clock_symbol = symbol

                # Check GHR / PHR
                if full_path == "TOP.Core.frontend.bpu.tage.io_req_ghr" or full_path == "TOP.Core.frontend.bpu.tage_io_req_ghr":
                    ghr_symbol = symbol
                if full_path == "TOP.Core.frontend.bpu.ittage.io_req_phr" or full_path == "TOP.Core.frontend.bpu.ittage_io_req_phr":
                    phr_symbol = symbol

                # Regs
                m_reg = re.match(r'debug_regs_(\d+)', name)
                if m_reg:
                    idx = int(m_reg.group(1))
                    preg_symbols[idx] = symbol

                # TAGE signals
                for t in range(4):
                    prefix_dot = f"TOP.Core.frontend.bpu.tage.tables_{t}."
                    prefix_under = f"TOP.Core.frontend.bpu.tage.tables_{t}_"
                    if full_path.startswith(prefix_dot) or full_path.startswith(prefix_under):
                        if name.endswith("tags_MPORT_en"): tage_symbols[t]['tag_en'].append(symbol)
                        elif name.endswith("tags_MPORT_addr"): tage_symbols[t]['tag_addr'].append(symbol)
                        elif name.endswith("tags_MPORT_data"): tage_symbols[t]['tag_data'].append(symbol)
                        elif name.endswith("ctrs_MPORT_en") or name.endswith("ctrs_MPORT_1_en"): tage_symbols[t]['ctr_en'].append(symbol)
                        elif name.endswith("ctrs_MPORT_addr") or name.endswith("ctrs_MPORT_1_addr"): tage_symbols[t]['ctr_addr'].append(symbol)
                        elif name.endswith("ctrs_MPORT_data") or name.endswith("ctrs_MPORT_1_data"): tage_symbols[t]['ctr_data'].append(symbol)
                        elif "us_MPORT_" in name:
                            m_p = re.search(r'us_MPORT_(\d+)', name)
                            if m_p:
                                p = int(m_p.group(1))
                                if p not in tage_symbols[t]['us']: tage_symbols[t]['us'][p] = {'en': [], 'addr': [], 'data': []}
                                if name.endswith("_en") or name.endswith(".en"): tage_symbols[t]['us'][p]['en'].append(symbol)
                                elif name.endswith("_addr") or name.endswith(".addr"): tage_symbols[t]['us'][p]['addr'].append(symbol)
                                elif name.endswith("_data") or name.endswith(".data"): tage_symbols[t]['us'][p]['data'].append(symbol)
                        elif name.endswith("io_update_valid"): tage_symbols[t]['update_valid'].append(symbol)
                        elif name.endswith("io_allocate"): tage_symbols[t]['allocate'].append(symbol)
                        elif name.endswith("io_hit"): tage_symbols[t]['hit'].append(symbol)
                        elif name.endswith("tags_read_tag_addr") or name.endswith("read_tag_addr"): tage_symbols[t]['read_tag_addr'].append(symbol)
                        elif name.endswith("req_tag"): tage_symbols[t]['req_tag'].append(symbol)

                # ITTAGE signals
                for t in range(4):
                    prefix_dot = f"TOP.Core.frontend.bpu.ittage.tables_{t}."
                    prefix_under = f"TOP.Core.frontend.bpu.ittage.tables_{t}_"
                    if full_path.startswith(prefix_dot) or full_path.startswith(prefix_under):
                        if name.endswith("tags_MPORT_en"): ittage_symbols[t]['tag_en'].append(symbol)
                        elif name.endswith("tags_MPORT_addr"): ittage_symbols[t]['tag_addr'].append(symbol)
                        elif name.endswith("tags_MPORT_data"): ittage_symbols[t]['tag_data'].append(symbol)
                        elif name.endswith("targets_MPORT_en") or name.endswith("targets_MPORT_1_en"): ittage_symbols[t]['tgt_en'].append(symbol)
                        elif name.endswith("targets_MPORT_addr") or name.endswith("targets_MPORT_1_addr"): ittage_symbols[t]['tgt_addr'].append(symbol)
                        elif name.endswith("targets_MPORT_data") or name.endswith("targets_MPORT_1_data"): ittage_symbols[t]['tgt_data'].append(symbol)
                        elif "us_MPORT_" in name:
                            m_p = re.search(r'us_MPORT_(\d+)', name)
                            if m_p:
                                p = int(m_p.group(1))
                                if p not in ittage_symbols[t]['us']: ittage_symbols[t]['us'][p] = {'en': [], 'addr': [], 'data': []}
                                if name.endswith("_en") or name.endswith(".en"): ittage_symbols[t]['us'][p]['en'].append(symbol)
                                elif name.endswith("_addr") or name.endswith(".addr"): ittage_symbols[t]['us'][p]['addr'].append(symbol)
                                elif name.endswith("_data") or name.endswith(".data"): ittage_symbols[t]['us'][p]['data'].append(symbol)
                        elif name.endswith("io_update_valid"): ittage_symbols[t]['update_valid'].append(symbol)
                        elif name.endswith("io_allocate"): ittage_symbols[t]['allocate'].append(symbol)
                        elif name.endswith("io_hit"): ittage_symbols[t]['hit'].append(symbol)
                        elif name.endswith("tags_read_tag_addr") or name.endswith("read_tag_addr"): ittage_symbols[t]['read_tag_addr'].append(symbol)
                        elif name.endswith("req_tag"): ittage_symbols[t]['req_tag'].append(symbol)

            elif stripped.startswith('$enddefinitions'):
                break

    print(f"Header Indexing complete. Clock Symbol: {clock_symbol}. GHR Symbol: {ghr_symbol}.")
    
    # Helper to parse binary VCD value
    def parse_bin(val_str):
        if not val_str: return 0
        if val_str.startswith('b'): val_str = val_str[1:]
        try: return int(val_str, 2)
        except ValueError: return 0

    # Initialize TAGE / ITTAGE Memory Models
    tage_mem = []
    for t in range(4):
        tage_mem.append({
            'tags': [0] * 128,
            'ctrs': [3] * 128,
            'us': [0] * 128,
            'valids': [False] * 128
        })

    ittage_mem = []
    for t in range(4):
        ittage_mem.append({
            'tags': [0] * 64,
            'targets': [0] * 64,
            'us': [0] * 64,
            'valids': [False] * 64
        })

    symbol_values = {}
    prev_clock = '0'
    cycle_count = 0

    # cycle snapshots
    snapshots = {}

    print("Step 2: Simulating VCD to collect cycle snapshots...")
    with open(vcd_path, 'r') as f:
        # Skip header
        for line in f:
            if line.strip().startswith('$enddefinitions'):
                break

        for line in f:
            stripped = line.strip()
            if not stripped or stripped.startswith('#'):
                continue
            
            # Value change lines
            if stripped.startswith('b') or stripped.startswith('r'):
                parts = stripped.split()
                if len(parts) == 2:
                    val, sym = parts
                    symbol_values[sym] = val
            else:
                val = stripped[0]
                sym = stripped[1:]
                symbol_values[sym] = val

            # Check rising clock edge
            if clock_symbol and sym == clock_symbol:
                current_clock = val
                if prev_clock == '0' and current_clock == '1':
                    cycle_count += 1
                    
                    if cycle_count > 1000:
                        break  # Optimize to first 1000 cycles

                    # Update TAGE memories based on current cycle's write commands
                    for t in range(4):
                        syms = tage_symbols[t]
                        tag_en = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['tag_en'])
                        tag_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tag_addr']), 0)
                        tag_data = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tag_data']), 0)

                        ctr_en = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['ctr_en'])
                        ctr_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['ctr_addr']), 0)
                        ctr_data = next((parse_bin(symbol_values.get(s, '0')) for s in syms['ctr_data']), 0)

                        us_en = 0
                        us_addr = 0
                        us_data = 0
                        for port, port_syms in syms['us'].items():
                            if any(parse_bin(symbol_values.get(s, '0')) == 1 for s in port_syms['en']):
                                us_en = 1
                                us_addr = next((parse_bin(symbol_values.get(s, '0')) for s in port_syms['addr']), 0)
                                us_data = next((parse_bin(symbol_values.get(s, '0')) for s in port_syms['data']), 0)
                                break

                        update_valid = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['update_valid'])
                        allocate = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['allocate'])

                        if tag_en: tage_mem[t]['tags'][tag_addr] = tag_data
                        if ctr_en: tage_mem[t]['ctrs'][ctr_addr] = ctr_data
                        if us_en:  tage_mem[t]['us'][us_addr] = us_data
                        if update_valid and allocate:
                            u_idx = tag_addr if tag_en else (ctr_addr if ctr_en else 0)
                            tage_mem[t]['valids'][u_idx] = True

                    # Update ITTAGE memories
                    for t in range(4):
                        syms = ittage_symbols[t]
                        tag_en = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['tag_en'])
                        tag_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tag_addr']), 0)
                        tag_data = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tag_data']), 0)

                        tgt_en = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['tgt_en'])
                        tgt_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tgt_addr']), 0)
                        tgt_data = next((parse_bin(symbol_values.get(s, '0')) for s in syms['tgt_data']), 0)

                        us_en = 0
                        us_addr = 0
                        us_data = 0
                        for port, port_syms in syms['us'].items():
                            if any(parse_bin(symbol_values.get(s, '0')) == 1 for s in port_syms['en']):
                                us_en = 1
                                us_addr = next((parse_bin(symbol_values.get(s, '0')) for s in port_syms['addr']), 0)
                                us_data = next((parse_bin(symbol_values.get(s, '0')) for s in port_syms['data']), 0)
                                break

                        update_valid = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['update_valid'])
                        allocate = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['allocate'])

                        if tag_en: ittage_mem[t]['tags'][tag_addr] = tag_data
                        if tgt_en: ittage_mem[t]['targets'][tgt_addr] = tgt_data
                        if us_en:  ittage_mem[t]['us'][us_addr] = us_data
                        if update_valid and allocate:
                            u_idx = tag_addr if tag_en else (tgt_addr if tgt_en else 0)
                            ittage_mem[t]['valids'][u_idx] = True

                    # Take Snapshot
                    snapshot = {
                        'ghr': parse_bin(symbol_values.get(ghr_symbol, '0')),
                        'phr': parse_bin(symbol_values.get(phr_symbol, '0')),
                        'pregs': {i: parse_bin(symbol_values.get(preg_symbols[i], '0')) for i in preg_symbols},
                        'tage_predictions': [],
                        'ittage_predictions': []
                    }

                    # Capture TAGE table details
                    for t in range(4):
                        syms = tage_symbols[t]
                        hit = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['hit'])
                        read_tag_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['read_tag_addr']), 0)
                        req_tag = next((parse_bin(symbol_values.get(s, '0')) for s in syms['req_tag']), 0)
                        
                        snapshot['tage_predictions'].append({
                            'hit': hit,
                            'idx': read_tag_addr,
                            'req_tag': req_tag,
                            'u': tage_mem[t]['us'][read_tag_addr],
                            'ctr': tage_mem[t]['ctrs'][read_tag_addr]
                        })

                    # Capture ITTAGE table details
                    for t in range(4):
                        syms = ittage_symbols[t]
                        hit = any(parse_bin(symbol_values.get(s, '0')) == 1 for s in syms['hit'])
                        read_tag_addr = next((parse_bin(symbol_values.get(s, '0')) for s in syms['read_tag_addr']), 0)
                        req_tag = next((parse_bin(symbol_values.get(s, '0')) for s in syms['req_tag']), 0)
                        
                        snapshot['ittage_predictions'].append({
                            'hit': hit,
                            'idx': read_tag_addr,
                            'req_tag': req_tag,
                            'u': ittage_mem[t]['us'][read_tag_addr],
                            'target': ittage_mem[t]['targets'][read_tag_addr]
                        })

                    snapshots[cycle_count] = snapshot

                prev_clock = current_clock

    print(f"Simulation parsing complete. Captured {len(snapshots)} cycle snapshots.")

    print("Step 3: Parsing sim_output2.txt for execution trace...")
    current_sim_cycle = 0
    current_cpu_cycle = 0

    all_renamed = []
    last_redirect_pc = None
    next_redirect_idx = 0
    next_bru_idx = 0
    ftb = {0: "EMPTY", 1: "EMPTY"}

    with open(log_path, 'r') as f:
        for line in f:
            m = re.match(r'\[TESTBENCH\] Cycle (\d+)', line)
            if m:
                current_sim_cycle = int(m.group(1))
                current_cpu_cycle = current_sim_cycle - 5
                continue

            m = re.search(r'\[FTB UPDATE\] pc=([0-9a-fA-F]+) index=\s*(\d+).*target=([0-9a-fA-F]+)', line)
            if m:
                src = int(m.group(1), 16) & 0xff
                idx = int(m.group(2))
                tgt = int(m.group(3), 16) & 0xff
                ftb[idx] = f"x{src:02x} - x{tgt:02x}"
                continue

            m = re.search(r'CORE RENAME \[Cycle\s+(\d+)\].*pc=([0-9a-fA-F]+).*inst=([0-9a-fA-F]+)', line)
            if m:
                rename_cpu_cycle = int(m.group(1))
                pc_val = int(m.group(2), 16)
                inst_hex = m.group(3)
                
                # Filter out alignment NOPs
                if (pc_val & 0x3) != 0:
                    continue

                lrd = None
                pdest = None
                m_lrd = re.search(r'lrd=\s*(\d+)', line)
                m_pdest = re.search(r'pdest=\s*(\d+)', line)
                if m_lrd: lrd = int(m_lrd.group(1))
                if m_pdest: pdest = int(m_pdest.group(1))

                pc_offset = pc_val & 0xff
                inst_text = disasm.get(pc_offset, f"unknown (0x{inst_hex})")

                inst_dict = {
                    'pc_val': pc_val,
                    'pc_short': f"x{pc_offset:02x}",
                    'instruction': inst_text,
                    'lrd': lrd,
                    'pdest': pdest,
                    'rename_cycle': rename_cpu_cycle,
                    'commit_cycle': None,
                    'bru_cycle': None,
                    'flushed': False,
                    'ftb0': ftb[0],
                    'ftb1': ftb[1]
                }
                all_renamed.append(inst_dict)
                continue

            m = re.search(r'\[BRU REDIRECT\] pc=([0-9a-fA-F]+)', line)
            if m:
                pc_val = int(m.group(1), 16)
                last_redirect_pc = pc_val
                for idx in range(next_bru_idx, len(all_renamed)):
                    inst = all_renamed[idx]
                    if inst['pc_val'] == pc_val and inst['bru_cycle'] is None and not inst['flushed']:
                        inst['bru_cycle'] = current_cpu_cycle
                        next_bru_idx = idx + 1
                        break
                continue

            m = re.search(r'\[REGFILE WRITE Port \d+\]: addr=\s*(\d+).*at cycle=\s*(\d+)', line)
            if m:
                pReg = int(m.group(1))
                write_cycle = int(m.group(2))
                for inst in all_renamed:
                    if inst['pdest'] == pReg and inst['commit_cycle'] is None and not inst['flushed']:
                        inst['commit_cycle'] = write_cycle
                        break
                continue

            m = re.search(r'\[BPU GHR RESTORE\].*redirect_pc=([0-9a-fA-F]+)', line)
            if m:
                redirect_pc = int(m.group(1), 16)
                idx_redirect = -1
                for idx in range(next_redirect_idx, len(all_renamed)):
                    if all_renamed[idx]['pc_val'] == redirect_pc and not all_renamed[idx]['flushed']:
                        idx_redirect = idx
                        break
                if idx_redirect != -1:
                    for idx in range(idx_redirect + 1, len(all_renamed)):
                        all_renamed[idx]['flushed'] = True
                    next_redirect_idx = idx_redirect + 1
                continue

    return snapshots, all_renamed

def generate_reports():
    snapshots, all_renamed = parse_vcd_and_log()
    if not snapshots or not all_renamed:
        return

    print("Step 4: Aligning snapshots and mapping register states using architectural sequence...")
    
    arch_trace = []
    pc = 0x80000000
    regs = {1: 0, 4: 0, 5: 0, 14: 0, 15: 0, 17: 0}
    
    # Execute program architecturally to generate candidates
    for _ in range(200):
        pc_short = pc & 0xff
        inst_text = disasm.get(pc_short, f"unknown (0x{pc:X})")
        
        if pc_short == 0x00:
            regs[1] = 10
            pc += 4
        elif pc_short == 0x04:
            regs[5] = 0
            pc += 4
        elif pc_short == 0x08:
            regs[5] += 6
            pc += 4
        elif pc_short == 0x0c:
            regs[14] = regs[5] & 3
            pc += 4
        elif pc_short == 0x10:
            if regs[14] == 0:
                pc += 8
            else:
                pc += 4
        elif pc_short == 0x14:
            regs[15] = 1
            pc += 4
        elif pc_short == 0x18:
            regs[17] = regs[14] << 2
            pc += 4
        elif pc_short == 0x1c:
            regs[4] = pc + 4
            pc = 0x80000030
        elif pc_short == 0x20:
            regs[15] = 10
            pc += 4
        elif pc_short == 0x24:
            pc = 0x8000003c
        elif pc_short == 0x28:
            regs[15] = 20
            pc += 4
        elif pc_short == 0x2c:
            pc = 0x8000003c
        elif pc_short == 0x30:
            regs[4] = regs[4] + regs[17]
            pc += 4
        elif pc_short == 0x34:
            temp = pc + 4
            pc = regs[4]
            regs[1] = temp
        elif pc_short == 0x38:
            regs[1] -= 1
            pc += 4
        elif pc_short == 0x3c:
            if regs[1] != 0:
                pc = 0x80000008
            else:
                pc += 4
        else:
            break
            
        arch_trace.append({
            'pc_val': pc_short,
            'instruction': inst_text,
            'regs': dict(regs)
        })

    # Now match with all_renamed
    final_rows = []
    matched_rename_indices = set()
    order_num = 1
    
    for arch_inst in arch_trace:
        matched_inst = None
        for idx, inst in enumerate(all_renamed):
            if idx in matched_rename_indices:
                continue
            if inst['flushed']:
                continue
            if (inst['pc_val'] & 0xff) == arch_inst['pc_val']:
                matched_inst = inst
                matched_rename_indices.add(idx)
                break
                
        if matched_inst is None:
            break
            
        rename_sim_cycle = matched_inst['rename_cycle'] + 5
        snap_ren = snapshots.get(rename_sim_cycle, None)
        
        reg_vals = {}
        for logical_reg in [1, 4, 5, 14, 15, 17]:
            val = arch_inst['regs'][logical_reg]
            reg_vals[logical_reg] = f"0x{val:X}" if val > 9 else str(val)
            
        ghr_val = snap_ren['ghr'] if snap_ren else 0
        phr_val = snap_ren['phr'] if snap_ren else 0
        
        tage_info = "-"
        if "beq" in matched_inst['instruction'] or "bne" in matched_inst['instruction']:
            provider_table = -1
            if snap_ren:
                for t in reversed(range(4)):
                    if snap_ren['tage_predictions'][t]['hit']:
                        provider_table = t
                        break
            if provider_table != -1:
                pred = snap_ren['tage_predictions'][provider_table]
                tage_info = f"T{provider_table}[{pred['idx']}], Tag=0x{pred['req_tag']:02X}, US={pred['u']}, CTR={pred['ctr']}"
            else:
                tage_info = "Bimodal Fallback"

        ittage_info = "-"
        if "jalr" in matched_inst['instruction']:
            provider_table = -1
            if snap_ren:
                for t in reversed(range(4)):
                    if snap_ren['ittage_predictions'][t]['hit']:
                        provider_table = t
                        break
            if provider_table != -1:
                pred = snap_ren['ittage_predictions'][provider_table]
                ittage_info = f"T{provider_table}[{pred['idx']}], Tag=0x{pred['req_tag']:02X}, Target=0x{pred['target']:08X}"
            else:
                ittage_info = "Indirect Fallback"

        row = {
            'order': str(order_num),
            'pc': matched_inst['pc_short'],
            'instruction': matched_inst['instruction'],
            'commit_cycle': str(matched_inst['commit_cycle']) if matched_inst['commit_cycle'] else "-",
            'bru_cycle': str(matched_inst['bru_cycle']) if matched_inst['bru_cycle'] else "-",
            'x1': reg_vals[1],
            'x4': reg_vals[4],
            'x5': reg_vals[5],
            'x14': reg_vals[14],
            'x15': reg_vals[15],
            'x17': reg_vals[17],
            'ftb0': matched_inst['ftb0'],
            'ftb1': matched_inst['ftb1'],
            'ghr': f"0b{ghr_val:b}" if ghr_val else "0",
            'phr': f"0b{phr_val:b}" if phr_val else "0",
            'tage': tage_info,
            'ittage': ittage_info,
            'is_spec': False,
            'is_br': "beq" in matched_inst['instruction'] or "bne" in matched_inst['instruction'] or "jal" in matched_inst['instruction']
        }
        final_rows.append(row)
        order_num += 1

    print("Step 5: Writing Styled Excel file (tage_test_trace_gemini.xlsx)...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Zaqal BPU Trace"

    headers = [
        "Order", "PC", "Instruction", "Cycle Commit PRF", "Cycle BRU",
        "x1", "x4", "x5", "x14", "x15", "x17",
        "FTB Entry 0 (Src-Tgt)", "FTB Entry 1 (Src-Tgt)", "GHR (TAGE index)", "TAGE Details",
        "PHR (ITTAGE index)", "ITTAGE Details"
    ]
    ws.append(headers)

    # Style definitions
    font_normal = Font(name="Consolas", size=10, color="FFFF00") # Yellow text
    font_spec = Font(name="Consolas", size=10, color="DDA0DD")   # Purple text
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF") # White header
    
    fill_header = PatternFill(start_color="202020", end_color="202020", fill_type="solid")
    fill_row_normal = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_row_br = PatternFill(start_color="332B00", end_color="332B00", fill_type="solid") # Dark gold background

    thin_border = Border(
        left=Side(style='thin', color='555555'),
        right=Side(style='thin', color='555555'),
        top=Side(style='thin', color='555555'),
        bottom=Side(style='thin', color='555555')
    )

    # Header styling
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row styling
    for r_idx, row_data in enumerate(final_rows, start=2):
        row_values = [
            row_data['order'], row_data['pc'], row_data['instruction'],
            row_data['commit_cycle'], row_data['bru_cycle'],
            row_data['x1'], row_data['x4'], row_data['x5'], row_data['x14'], row_data['x15'], row_data['x17'],
            row_data['ftb0'], row_data['ftb1'], row_data['ghr'], row_data['tage'],
            row_data['phr'], row_data['ittage']
        ]
        ws.append(row_values)
        
        # Apply style to each cell
        is_br = row_data['is_br']
        is_spec = row_data['is_spec']
        for col_idx in range(1, len(row_values) + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.border = thin_border
            cell.font = font_spec if is_spec else font_normal
            cell.fill = fill_row_br if (is_br and not is_spec) else fill_row_normal
            cell.alignment = Alignment(horizontal="left" if col_idx == 3 else "center", vertical="center")

    # Save
    excel_out_path = 'programs/assembly/tage_test_trace_gemini.xlsx'
    wb.save(excel_out_path)
    print(f"Excel report saved to {excel_out_path}.")

    print("Step 6: Writing Styled HTML dashboard (tage_test_trace_gemini.html)...")
    # Generate interactive HTML report
    html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Zaqal Branch Prediction Unit Trace Dashboard</title>
    <style>
        body {
            background-color: #0d0d0d;
            color: #e0e0e0;
            font-family: 'Consolas', 'Courier New', monospace;
            margin: 20px;
        }
        h1 {
            color: #FFD700;
            text-align: center;
            border-bottom: 2px solid #333;
            padding-bottom: 10px;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        .table-container {
            overflow-x: auto;
            margin-top: 20px;
            border-radius: 8px;
            border: 1px solid #333;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            background-color: #000000;
        }
        th {
            background-color: #202020;
            color: #ffffff;
            font-weight: bold;
            padding: 12px;
            border: 1px solid #333;
            font-size: 13px;
        }
        td {
            padding: 8px 12px;
            border: 1px solid #333;
            font-size: 12px;
            color: #FFFF00; /* Default yellow */
        }
        tr.branch-row {
            background-color: #332B00; /* Dark gold fill */
        }
        tr.speculative-row td {
            color: #DDA0DD !important; /* Speculative purple text */
        }
        tr:hover {
            background-color: #1a1a1a;
        }
        .search-box {
            margin-bottom: 20px;
            padding: 10px;
            width: 100%;
            max-width: 400px;
            background-color: #151515;
            border: 1px solid #444;
            color: #fff;
            font-family: inherit;
            font-size: 14px;
            border-radius: 4px;
        }
    </style>
</head>
<body>
    <h1>Zaqal BPU / TAGE Trace Dashboard (First 1000 Cycles)</h1>
    <input type="text" id="search" class="search-box" placeholder="Search instruction, PC, cycle..." onkeyup="filterTable()">
    
    <div class="table-container">
        <table id="trace-table">
            <thead>
                <tr>
                    <th>Order</th>
                    <th>PC</th>
                    <th>Instruction</th>
                    <th>Cycle Commit PRF</th>
                    <th>Cycle BRU</th>
                    <th>x1</th>
                    <th>x4</th>
                    <th>x5</th>
                    <th>x14</th>
                    <th>x15</th>
                    <th>x17</th>
                    <th>FTB Entry 0</th>
                    <th>FTB Entry 1</th>
                    <th>GHR</th>
                    <th>TAGE Details</th>
                    <th>PHR</th>
                    <th>ITTAGE Details</th>
                </tr>
            </thead>
            <tbody>
"""
    for row in final_rows:
        row_class = ""
        if row['is_spec']:
            row_class = 'class="speculative-row"'
        elif row['is_br']:
            row_class = 'class="branch-row"'
            
        html_content += f"""                <tr {row_class}>
                    <td>{row['order']}</td>
                    <td>{row['pc']}</td>
                    <td style="text-align: left;">{row['instruction']}</td>
                    <td>{row['commit_cycle']}</td>
                    <td>{row['bru_cycle']}</td>
                    <td>{row['x1']}</td>
                    <td>{row['x4']}</td>
                    <td>{row['x5']}</td>
                    <td>{row['x14']}</td>
                    <td>{row['x15']}</td>
                    <td>{row['x17']}</td>
                    <td>{row['ftb0']}</td>
                    <td>{row['ftb1']}</td>
                    <td>{row['ghr']}</td>
                    <td>{row['tage']}</td>
                    <td>{row['phr']}</td>
                    <td>{row['ittage']}</td>
                </tr>\n"""

    html_content += """            </tbody>
        </table>
    </div>

    <script>
        function filterTable() {
            var input = document.getElementById("search");
            var filter = input.value.toUpperCase();
            var table = document.getElementById("trace-table");
            var tr = table.getElementsByTagName("tr");

            for (var i = 1; i < tr.length; i++) {
                var show = false;
                var tds = tr[i].getElementsByTagName("td");
                for (var j = 0; j < tds.length; j++) {
                    if (tds[j]) {
                        var text = tds[j].textContent || tds[j].innerText;
                        if (text.toUpperCase().indexOf(filter) > -1) {
                            show = true;
                            break;
                        }
                    }
                }
                tr[i].style.display = show ? "" : "none";
            }
        }
    </script>
</body>
</html>
"""
    html_out_path = 'programs/assembly/tage_test_trace_gemini.html'
    with open(html_out_path, 'w') as f:
        f.write(html_content)
    print(f"HTML dashboard saved to {html_out_path}.")

if __name__ == '__main__':
    generate_reports()
